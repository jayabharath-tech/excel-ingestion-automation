"""Write a pandas DataFrame to a clean Excel file."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from core.recipe_engine import Recipe

from core.target_schema import TABLE_SPEC

logger = logging.getLogger(__name__)


def _recipe_to_dataframe(recipe: Recipe) -> pd.DataFrame:
    """Convert a Recipe into a flat DataFrame for the metadata sheet."""
    rows = []
    for t in recipe.transformations:
        rule_str = json.dumps(t.rule) if t.rule else ""
        alt_str = ", ".join(t.alternatives) if t.alternatives else ""
        rows.append({
            "Target Column": t.target,
            "Type": t.type.value,
            "Source": (
                ", ".join(t.source) if isinstance(t.source, list) else (t.source or "")
            ),
            "Rule": rule_str,
            "Target Type": t.target_type or "",
            "Confidence": t.confidence if t.confidence is not None else "",
            "Alternatives": alt_str,
        })
    return pd.DataFrame(rows)


def _apply_excel_formatting(output_path: Path) -> None:
    """
    Apply Excel column formatting based on TABLE_SPEC schema.
    Safe, fast, and production-ready.
    """

    if not output_path.exists():
        logger.warning(f"Output file does not exist: {output_path}")
        return

    try:
        wb = load_workbook(str(output_path))

        if "Use" not in wb.sheetnames:
            logger.warning(f"Data sheet not found. Available sheets: {wb.sheetnames}")
            return

        ws = wb["Use"]

        # Map column names -> column index
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        col_name_to_idx = {
            name: idx + 1
            for idx, name in enumerate(header_row)
            if name
        }

        logger.debug(f"Detected columns: {list(col_name_to_idx.keys())}")

        formatted_columns = 0

        for col in TABLE_SPEC.columns:

            # Skip columns without formatting rules
            if not col.excel_format:
                continue

            if col.name not in col_name_to_idx:
                logger.warning(f"Column '{col.name}' not found in Excel")
                continue

            col_idx = col_name_to_idx[col.name]
            column_letter = get_column_letter(col_idx)

            column_cells = next(ws.iter_cols(
                min_col=col_idx,
                max_col=col_idx,
                min_row=2,
                max_row=ws.max_row
            ))

            formatted_cells = 0

            for cell in column_cells:

                if cell.value is None:
                    continue

                # Ensure text columns stay text (e.g., phone numbers)
                if col.excel_format == "@":
                    cell.value = str(cell.value)

                cell.number_format = col.excel_format
                formatted_cells += 1

            logger.info(
                f"Formatted column '{col.name}' ({column_letter}) "
                f"with '{col.excel_format}' on {formatted_cells} cells"
            )

            formatted_columns += 1

        wb.save(str(output_path))

        logger.info(
            f"Excel formatting completed successfully. "
            f"{formatted_columns} columns formatted."
        )

    except Exception as e:
        logger.error(f"Excel formatting failed: {e}")
        import traceback
        logger.error(traceback.format_exc())


def write_excel(
    df: pd.DataFrame,
    output_path: str | Path,
    sheet_name: str = "Use",
    recipe: Recipe | None = None,
    apply_formatting: bool = True,
) -> Path:
    """Write *df* to an .xlsx file.

    If *recipe* is provided, a second sheet named "Transformations" is added
    with the column mapping details.

    Creates parent directories if they don't exist.  Returns the resolved
    output path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Debug: Check DataFrame columns
    logger.debug(f"DataFrame columns: {list(df.columns)}")
    logger.debug(f"DataFrame dtypes: {dict(df.dtypes)}")
    logger.debug(f"DataFrame shape: {df.shape}")
    
    # Ensure DataFrame has a clean index (no named index)
    df_to_write = df.reset_index(drop=True)
    
    # Convert IdNo column to string to prevent exponential formatting in Excel
    if "IdNo" in df_to_write.columns:
        df_to_write["IdNo"] = df_to_write["IdNo"].astype(str)
    
    logger.info(f"Writing DataFrame with shape {df_to_write.shape} to {output_path}")

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)

        if recipe is not None:
            recipe_df = _recipe_to_dataframe(recipe)
            recipe_df.to_excel(
                writer, sheet_name="Transformations", index=False
            )

    # Apply Excel formatting only if requested
    if apply_formatting:
        # _apply_excel_formatting(output_path)
        _apply_excel_formatting(output_path)

    return output_path